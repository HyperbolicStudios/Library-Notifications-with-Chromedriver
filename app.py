#Google Calendar Modules

from apiclient import discovery
from apiclient.discovery import build
from httplib2 import Http
from oauth2client import file, client, tools

#Selenium Modules
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options

#spreadsheet
import openpyxl

#os Modules
import os
from inspect import getsourcefile
from os.path import abspath

#General Modules
import time
import datetime
import traceback
import pytz


from datetime import datetime
from pytz import timezone

#ID location of the form - do not change
ID_A = "j_username"
ID_B = "j_password"

date_format='%m/%d/%Y %H:%M:%S %Z'
date = datetime.now(tz=pytz.utc)
date = date.astimezone(timezone('US/Pacific'))
date  = (date.strftime(date_format))

error_message = "Program started at " + date
#Your username and password
username = "369135m"
password = "MashedPotatoes"


#Log in to google API
SCOPES = 'https://www.googleapis.com/auth/calendar'
store = file.Storage('storage.json')
creds = store.get()
if not creds or creds.invalid:
    flow = client.flow_from_clientsecrets('client_secret.json', SCOPES)
    creds = tools.run_flow(flow, store)
GCAL = discovery.build('calendar', 'v3', http=creds.authorize(Http()))

#Find the curent file location/chdir to current location
directory = abspath(getsourcefile(lambda:0))
newDirectory = directory[:(directory.rfind("\\")+1)]
os.chdir(newDirectory)


#Calendar API Setup

def calendar (title, date):


    GMT_OFF = '-07:00'      # PDT/MST/GMT-7

    EVENT = { #Event description
        'summary': 'GVPL Item Due',
        'start':  {'dateTime': date+'T06:00:00%s' % GMT_OFF},
        'end':    {'dateTime': date+'T06:30:00%s' % GMT_OFF},
        'description': ('Item: ' + title)

    }

    e = GCAL.events().insert(calendarId='primary',
            sendNotifications=True, body=EVENT).execute()

    print("Event added!")


def get_date(text):
    length = len(text)
    date = text[(length-9):length]

    length = len(date)

    year = "20" + (date[(length-2):length])


    date = (date[:(length-3)])

    location = date.find("/")

    day = (date[(location+1):(location+3)])
    month = (date[(location-2):(location)])


    #Format to yyyy-mm-dd

    date = year + "-" + month.replace(" ", "") + "-" + day
    return date

    #-----------------------------------------

def searchForItem(data):

    foundItem = False
    i2 = sheet.max_row
    while(i2 > 0):
        loc = ('A' + str(i2))


        if (sheet[loc].value == data[0] and sheet[('B'+str(i2))].value == data[1] and sheet[('C' + str(i2))].value == data[2] and str(sheet[('D'+ str(i2))].value) == data[3]):
            print("Found!")
            foundItem = True
            i2 = -0

        i2 = i2-1

    return(foundItem)







try:
    i =1;


    #Chromedriver options
    options = Options()
    options.add_argument('--headless') #Comment out to see browser window in real-time

    PATH = ("chromedriver.exe") #Assumes chromedriver is in the same folder as the program


    driver = webdriver.Chrome(chrome_options=options, executable_path=PATH)
    time.sleep(1)
    driver.get("""https://gvpl.ent.sirsidynix.net/client/en_US/default/search/patronlogin/https:$002f$002fgvpl.ent.sirsidynix.net$002fclient$002fen_US$002fdefault$002fsearch$002faccount$003f/"""
    ) #navigate
    print("Starting.")

    time.sleep(4)
    #Login
    driver.find_element_by_id(ID_A).send_keys(username)
    driver.find_element_by_id(ID_B).send_keys(password)
    driver.find_element_by_id("submit_0").click()

    #User Page
    time.sleep(2)
    driver.find_element_by_link_text("Checkouts").click()
    el=(driver.find_element_by_css_selector("html"))

    text = el.text
    #Get number of items
    x = text.find("Total Items Checked Out:")
    x = (text[(x+25):(x+27)])
    x = int(x) + 1
    while i < x: #Loop through each item
        argument = "//table[@id='myCheckouts_checkoutslist_table']/tbody/tr[@class='checkoutsLine'][" + (str(i)) + "]"
        el=(driver.find_element_by_xpath(argument))

        text = el.text

        #-------------------------
        #Get the due date using function defined above
        date = get_date(text)

        print(text)

        #Extract the title
        title = text.splitlines()[0]
        title1 = text.splitlines()[1]

        #Some items print out the title twice, for some reason. Check if this is so.
        if title == title1:
            line = 0
        else:
            line = 1
        author = text.splitlines()[2-line]
        ddcode = text.splitlines()[3-line]
        barcode = text.splitlines()[4-line]

        #This is the list of item info
        bookData = [title, author, ddcode, barcode]

        #Open excell
        wb = openpyxl.load_workbook('data.xlsx')
        sheet = wb.get_sheet_by_name('Sheet1')

        #Search for item using function defined above

        result = (searchForItem(bookData))
        print("----------------------------------------")

        if result == False: #Add to spreadsheet if item not found
            loc = str(sheet.max_row + 1)
            print("Adding to spreadsheet.")
            sheet[('A'+loc)] = bookData[0]
            sheet[('B'+loc)] = bookData[1]
            sheet[('C'+loc)] = bookData[2]
            sheet[('D'+loc)] = bookData[3]

            calendar(bookData[0], date) #Add to calendar using fuction defined above



        i = i+1

        wb.save("data.xlsx")
except:
    error_message = error_message + "\\n" + (traceback.format_exc()) + "\\n---------------"
    print(error_message)
driver.quit()

file = open('logs.txt', 'a')
file.write(error_message + "\\n")
file.close()

print("Done!")
